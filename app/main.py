from datetime import date, datetime
import json
import io
import os

from fastapi import FastAPI, Request, Form, Depends, HTTPException
from fastapi.responses import HTMLResponse, RedirectResponse, StreamingResponse, FileResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from sqlalchemy.orm import Session
import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)

from app.database import SessionLocal, engine, Base
from app.models.client import Client, ContentItem
from app.services.ai_content import generate_content_calendar
from app.services.pricing import POST_TYPES, calculate_item_cost, calculate_calendar_cost

Base.metadata.create_all(bind=engine)

# Safe column migrations for new fields
from sqlalchemy import text as _text
with engine.connect() as _conn:
    for _col, _def in [
        ("client_photo_path", "VARCHAR DEFAULT ''"),
        ("creative_paths",    "TEXT DEFAULT '[]'"),
        ("posted_at",         "VARCHAR DEFAULT ''"),
        ("posted_to",         "TEXT DEFAULT '[]'"),
    ]:
        try:
            _conn.execute(_text(f"ALTER TABLE content_items ADD COLUMN {_col} {_def}"))
            _conn.commit()
        except Exception:
            pass

app = FastAPI(title="Influz Studio - System")
app.mount("/static", StaticFiles(directory="app/static"), name="static")
templates = Jinja2Templates(directory="app/templates")
templates.env.filters["fromjson"] = json.loads


@app.get("/creatives/{filename}")
def serve_creative(filename: str):
    """Serve generated creatives from /tmp (Render writable filesystem)."""
    path = Path(f"/tmp/creatives/{filename}")
    if path.exists():
        return FileResponse(str(path), media_type="image/png")
    raise HTTPException(status_code=404, detail="Creative not found")


def get_db():
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()


@app.get("/", response_class=HTMLResponse)
def home(request: Request, db: Session = Depends(get_db)):
    clients = db.query(Client).order_by(Client.created_at.desc()).all()
    return templates.TemplateResponse("index.html", {"request": request, "clients": clients})


@app.get("/onboard", response_class=HTMLResponse)
def onboard_form(request: Request):
    return templates.TemplateResponse("onboard.html", {"request": request})


@app.post("/onboard")
def onboard_submit(
    request: Request,
    business_name: str = Form(...),
    niche: str = Form(...),
    brand_voice: str = Form(...),
    goals: str = Form(...),
    city: str = Form(""),
    instagram_handle: str = Form(""),
    facebook_page: str = Form(""),
    notes: str = Form(""),
    posts_per_month: int = Form(16),
    rate_story: float = Form(150.0),
    rate_post: float = Form(300.0),
    rate_carousel: float = Form(500.0),
    rate_reel: float = Form(800.0),
    rate_ugc: float = Form(600.0),
    rate_on_demand: float = Form(1000.0),
    db: Session = Depends(get_db),
):
    client = Client(
        business_name=business_name,
        niche=niche,
        brand_voice=brand_voice,
        goals=goals,
        city=city,
        instagram_handle=instagram_handle,
        facebook_page=facebook_page,
        notes=notes,
        posts_per_month=posts_per_month,
        rate_story=rate_story,
        rate_post=rate_post,
        rate_carousel=rate_carousel,
        rate_reel=rate_reel,
        rate_ugc=rate_ugc,
        rate_on_demand=rate_on_demand,
    )
    db.add(client)
    db.commit()
    db.refresh(client)
    return RedirectResponse(url=f"/client/{client.id}", status_code=303)


@app.get("/client/{client_id}", response_class=HTMLResponse)
def client_detail(request: Request, client_id: int, db: Session = Depends(get_db)):
    client = db.query(Client).filter(Client.id == client_id).first()
    if not client:
        raise HTTPException(status_code=404, detail="Client not found")
    content_items = (
        db.query(ContentItem)
        .filter(ContentItem.client_id == client_id)
        .order_by(ContentItem.post_date)
        .all()
    )
    cost_summary = calculate_calendar_cost(content_items, client)
    item_costs = {item.id: calculate_item_cost(item, client) for item in content_items}
    months = build_calendar_months(content_items)
    return templates.TemplateResponse(
        "client_detail.html",
        {
            "request": request,
            "client": client,
            "content_items": content_items,
            "months": months,
            "cost_summary": cost_summary,
            "item_costs": item_costs,
            "post_types": POST_TYPES,
            "today": date.today().isoformat(),
        },
    )


def build_calendar_months(content_items: list[ContentItem]):
    if not content_items:
        return []
    items_by_date = {}
    for item in content_items:
        items_by_date.setdefault(item.post_date, []).append(item)
    start = min(items_by_date.keys())
    end = max(items_by_date.keys())
    months = []
    cursor = date(start.year, start.month, 1)
    end_month = date(end.year, end.month, 1)
    while cursor <= end_month:
        months.append((cursor.year, cursor.month))
        cursor = date(cursor.year + (cursor.month == 12), (cursor.month % 12) + 1, 1)
    month_data = []
    for year, month in months:
        first_day = date(year, month, 1)
        next_month_date = date(year + (month == 12), (month % 12) + 1, 1)
        days_in_month = (next_month_date - first_day).days
        leading_blanks = first_day.weekday()
        weeks, week = [], [None] * leading_blanks
        for day_num in range(1, days_in_month + 1):
            current = date(year, month, day_num)
            week.append({"date": current, "posts": items_by_date.get(current, [])})
            if len(week) == 7:
                weeks.append(week)
                week = []
        if week:
            while len(week) < 7:
                week.append(None)
            weeks.append(week)
        month_data.append({"label": first_day.strftime("%B %Y"), "weeks": weeks})
    return month_data


@app.post("/client/{client_id}/generate")
def generate_calendar(
    client_id: int,
    start_date: str = Form(None),
    num_days: int = Form(60),
    db: Session = Depends(get_db),
):
    client = db.query(Client).filter(Client.id == client_id).first()
    if not client:
        raise HTTPException(status_code=404, detail="Client not found")
    db.query(ContentItem).filter(
        ContentItem.client_id == client_id,
        ContentItem.status == "generated",
        ContentItem.is_on_demand == False,
    ).delete(synchronize_session=False)
    db.commit()
    start = datetime.strptime(start_date, "%Y-%m-%d").date() if start_date else date.today()
    num_posts = max(1, round(client.posts_per_month * num_days / 30))
    items = generate_content_calendar(
        business_name=client.business_name,
        niche=client.niche,
        brand_voice=client.brand_voice,
        goals=client.goals,
        start_date=start,
        num_days=num_days,
        num_posts=num_posts,
        city=client.city,
    )
    for item in items:
        db.add(ContentItem(
            client_id=client.id,
            post_date=datetime.strptime(item["post_date"], "%Y-%m-%d").date(),
            post_type=item.get("post_type", "Static"),
            platforms=json.dumps(item.get("platforms", ["instagram"])),
            topic=item.get("topic", ""),
            cover_text=item.get("cover_text", ""),
            image_text=item.get("image_text", ""),
            caption=item.get("caption", ""),
            hashtags=json.dumps([]),
            reference_note=item.get("reference_note", ""),
            status="generated",
            is_on_demand=False,
        ))
    db.commit()
    return RedirectResponse(url=f"/client/{client_id}", status_code=303)


@app.post("/client/{client_id}/add-on-demand")
def add_on_demand(
    client_id: int,
    post_date: str = Form(...),
    post_type: str = Form(...),
    platforms: list[str] = Form([]),
    topic: str = Form(""),
    cover_text: str = Form(""),
    caption: str = Form(""),
    db: Session = Depends(get_db),
):
    client = db.query(Client).filter(Client.id == client_id).first()
    if not client:
        raise HTTPException(status_code=404, detail="Client not found")
    db.add(ContentItem(
        client_id=client.id,
        post_date=datetime.strptime(post_date, "%Y-%m-%d").date(),
        post_type=post_type,
        platforms=json.dumps(platforms),
        topic=topic,
        cover_text=cover_text,
        caption=caption,
        hashtags=json.dumps([]),
        status="generated",
        is_on_demand=True,
    ))
    db.commit()
    return RedirectResponse(url=f"/client/{client_id}", status_code=303)


@app.post("/content/{item_id}/approve")
def approve_item(item_id: int, db: Session = Depends(get_db)):
    item = db.query(ContentItem).filter(ContentItem.id == item_id).first()
    if not item:
        raise HTTPException(status_code=404, detail="Not found")
    item.status = "approved"
    db.commit()
    return RedirectResponse(url=f"/client/{item.client_id}", status_code=303)


@app.post("/content/{item_id}/update")
def update_item(
    item_id: int,
    post_date: str = Form(...),
    post_type: str = Form(...),
    platforms: list[str] = Form([]),
    topic: str = Form(""),
    cover_text: str = Form(""),
    image_text: str = Form(""),
    caption: str = Form(""),
    reference_note: str = Form(""),
    client_feedback: str = Form(""),
    db: Session = Depends(get_db),
):
    item = db.query(ContentItem).filter(ContentItem.id == item_id).first()
    if not item:
        raise HTTPException(status_code=404, detail="Not found")
    item.post_date = datetime.strptime(post_date, "%Y-%m-%d").date()
    item.post_type = post_type
    item.platforms = json.dumps(platforms)
    item.topic = topic
    item.cover_text = cover_text
    item.image_text = image_text
    item.caption = caption
    item.reference_note = reference_note
    item.client_feedback = client_feedback
    db.commit()
    return RedirectResponse(url=f"/client/{item.client_id}", status_code=303)


@app.post("/content/{item_id}/delete")
def delete_item(item_id: int, db: Session = Depends(get_db)):
    item = db.query(ContentItem).filter(ContentItem.id == item_id).first()
    if not item:
        raise HTTPException(status_code=404, detail="Not found")
    client_id = item.client_id
    db.delete(item)
    db.commit()
    return RedirectResponse(url=f"/client/{client_id}", status_code=303)


@app.get("/client/{client_id}/export")
def export_calendar(client_id: int, db: Session = Depends(get_db)):
    """Export the content calendar as a professional branded Excel sheet."""
    client = db.query(Client).filter(Client.id == client_id).first()
    if not client:
        raise HTTPException(status_code=404, detail="Not found")
    items = (
        db.query(ContentItem)
        .filter(ContentItem.client_id == client_id)
        .order_by(ContentItem.post_date)
        .all()
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Content Calendar"

    # ── Styles ──────────────────────────────────────────────────────────────
    dark_fill    = PatternFill("solid", fgColor="0E1822")
    teal_fill    = PatternFill("solid", fgColor="2DD4BF")
    soft_fill    = PatternFill("solid", fgColor="13202E")
    amber_fill   = PatternFill("solid", fgColor="FBBF24")
    white_font   = Font(name="Calibri", color="F4F7F6", bold=False, size=11)
    dark_font    = Font(name="Calibri", color="0E1822", bold=True, size=11)
    teal_font    = Font(name="Calibri", color="2DD4BF", bold=True, size=11)
    muted_font   = Font(name="Calibri", color="7E8FA0", size=10)
    header_font  = Font(name="Calibri", color="0E1822", bold=True, size=11)
    title_font   = Font(name="Calibri", color="F4F7F6", bold=True, size=16)
    thin_border  = Border(
        bottom=Side(style="thin", color="2DD4BF"),
        top=Side(style="thin", color="2DD4BF"),
        left=Side(style="thin", color="13202E"),
        right=Side(style="thin", color="13202E"),
    )
    wrap = Alignment(wrap_text=True, vertical="top")

    # ── Title row ────────────────────────────────────────────────────────────
    ws.merge_cells("A1:K1")
    title_cell = ws["A1"]
    title_cell.value = f"Influz Studio  ·  {client.business_name}  ·  Content Calendar"
    title_cell.font = title_font
    title_cell.fill = dark_fill
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 36

    # Subtitle
    ws.merge_cells("A2:K2")
    sub_cell = ws["A2"]
    sub_cell.value = f"Crafting Digital Influence  ·  {client.posts_per_month} posts/month  ·  influzstudio@gmail.com"
    sub_cell.font = muted_font
    sub_cell.fill = dark_fill
    sub_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 20

    # Blank spacer
    ws.row_dimensions[3].height = 8
    for col in range(1, 12):
        ws.cell(row=3, column=col).fill = dark_fill

    # ── Headers ───────────────────────────────────────────────────────────────
    headers = [
        "#", "Date", "Day", "Post Type", "Platforms",
        "Topic", "Cover Text", "Image Text",
        "Caption / Hashtags", "Reference Note", "Client Feedback"
    ]
    col_widths = [4, 13, 10, 11, 18, 28, 28, 28, 55, 32, 24]

    for col_idx, (header, width) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=4, column=col_idx)
        cell.value = header
        cell.font = header_font
        cell.fill = teal_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border
        ws.column_dimensions[cell.column_letter].width = width
    ws.row_dimensions[4].height = 22

    # ── Data rows ─────────────────────────────────────────────────────────────
    post_type_colors = {
        "Reel":     "A78BFA",  # purple
        "Carousel": "34D399",  # green
        "Static":   "60A5FA",  # blue
        "Story":    "F9A8D4",  # pink
        "UGC":      "FCD34D",  # yellow
    }

    for row_idx, item in enumerate(items, start=5):
        platforms = ", ".join(p.capitalize() for p in json.loads(item.platforms or "[]"))
        day_name = item.post_date.strftime("%A")

        row_fill = soft_fill if not item.is_on_demand else PatternFill("solid", fgColor="1C1505")

        values = [
            row_idx - 4,
            item.post_date.strftime("%d-%b-%Y"),
            day_name,
            item.post_type,
            platforms,
            item.topic,
            item.cover_text,
            item.image_text,
            item.caption,
            item.reference_note,
            item.client_feedback or "",
        ]

        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = value
            cell.fill = row_fill
            cell.alignment = wrap
            cell.border = thin_border

            if col_idx == 1:  # post count
                cell.font = teal_font
                cell.alignment = Alignment(horizontal="center", vertical="top")
            elif col_idx == 3:  # day name
                cell.font = muted_font
            elif col_idx == 4:  # post type — colored badge feel
                ptype_color = post_type_colors.get(item.post_type, "7E8FA0")
                cell.font = Font(name="Calibri", color=ptype_color, bold=True, size=11)
                if item.is_on_demand:
                    cell.font = Font(name="Calibri", color="FBBF24", bold=True, size=11)
                    # Add ✦ marker
                    cell.value = f"✦ {item.post_type}"
            elif col_idx == 5:  # platforms
                cell.font = Font(name="Calibri", color="A7E8DC", size=10)
            elif col_idx in (6, 7, 8):  # topic/cover/image text
                cell.font = Font(name="Calibri", color="F4F7F6", bold=(col_idx == 6), size=11)
            elif col_idx == 9:  # caption
                cell.font = Font(name="Calibri", color="D1D5DB", size=10)
            elif col_idx == 10:  # reference
                cell.font = Font(name="Calibri", color="7E8FA0", size=10, italic=True)
            else:
                cell.font = white_font

        # Row height based on caption length
        caption_len = len(item.caption)
        ws.row_dimensions[row_idx].height = max(60, min(120, caption_len // 3))

    # ── Status key / footer ───────────────────────────────────────────────────
    footer_row = len(items) + 7
    ws.merge_cells(f"A{footer_row}:K{footer_row}")
    footer_cell = ws[f"A{footer_row}"]
    footer_cell.value = f"Generated by Influz Studio  ·  {datetime.today().strftime('%d %b %Y')}  ·  Crafting Digital Influence"
    footer_cell.font = Font(name="Calibri", color="2DD4BF", italic=True, size=10)
    footer_cell.fill = dark_fill
    footer_cell.alignment = Alignment(horizontal="center")

    # Freeze panes below header
    ws.freeze_panes = "A5"

    # ── Stream as download ────────────────────────────────────────────────────
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"Influz_Studio_{client.business_name.replace(' ', '_')}_Calendar.xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@app.get("/client/{client_id}/analytics", response_class=HTMLResponse)
def analytics(request: Request, client_id: int, db: Session = Depends(get_db)):
    client = db.query(Client).filter(Client.id == client_id).first()
    if not client:
        raise HTTPException(status_code=404, detail="Not found")
    sample_data = {
        "followers": 4820, "followers_growth": "+6.2%",
        "avg_reach": 2310, "reach_growth": "+11.4%",
        "engagement_rate": "4.8%", "engagement_growth": "+0.7pt",
        "posts_published": 12,
        "weekly_reach": [1450, 1820, 2090, 2310],
        "weekly_engagement": [58, 79, 95, 112],
    }
    return templates.TemplateResponse(
        "analytics.html", {"request": request, "client": client, "data": sample_data}
    )


# ── Excel import ──────────────────────────────────────────────────────────────
from fastapi import UploadFile, File
import openpyxl
import io as _io

@app.get("/client/{client_id}/import", response_class=HTMLResponse)
def import_form(request: Request, client_id: int, db: Session = Depends(get_db)):
    client = db.query(Client).filter(Client.id == client_id).first()
    if not client:
        raise HTTPException(status_code=404, detail="Client not found")
    return templates.TemplateResponse(
        "import_calendar.html", {"request": request, "client": client}
    )


@app.post("/client/{client_id}/import")
async def import_calendar(
    client_id: int,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
):
    client = db.query(Client).filter(Client.id == client_id).first()
    if not client:
        raise HTTPException(status_code=404, detail="Client not found")

    contents = await file.read()
    wb = openpyxl.load_workbook(_io.BytesIO(contents), data_only=True)
    ws = wb.active

    # Detect header row — look for a row containing "Date" in one of the first 10 rows
    header_row = None
    col_map = {}
    for row_idx in range(1, 11):
        for col_idx in range(1, 20):
            val = ws.cell(row_idx, col_idx).value
            if isinstance(val, str) and val.strip().lower() == "date":
                header_row = row_idx
                # Map column names → column indices
                for c in range(1, 20):
                    h = ws.cell(row_idx, c).value
                    if h:
                        col_map[str(h).strip().lower()] = c
                break
        if header_row:
            break

    if not header_row:
        raise HTTPException(status_code=400, detail="Could not find header row with 'Date' column")

    # Column aliases for flexibility
    def get_col(aliases: list[str]) -> int | None:
        for alias in aliases:
            if alias.lower() in col_map:
                return col_map[alias.lower()]
        return None

    col_date     = get_col(["date"])
    col_type     = get_col(["post type", "type"])
    col_platform = get_col(["platforms", "platform"])
    col_topic    = get_col(["topic"])
    col_cover    = get_col(["cover text", "cover"])
    col_image    = get_col(["image text", "image"])
    col_caption  = get_col(["caption / hashtags", "caption", "caption/hashtags"])
    col_ref      = get_col(["reference note", "reference"])

    if not col_date:
        raise HTTPException(status_code=400, detail="Date column not found in sheet")

    imported = 0
    errors = []

    for row_idx in range(header_row + 1, ws.max_row + 1):
        raw_date = ws.cell(row_idx, col_date).value if col_date else None
        if not raw_date:
            continue

        # Parse date — handles datetime objects or string formats
        post_date = None
        if hasattr(raw_date, "date"):
            post_date = raw_date.date()
        elif isinstance(raw_date, str):
            for fmt in ["%d-%b-%Y", "%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%d %b %Y"]:
                try:
                    post_date = datetime.strptime(raw_date.strip(), fmt).date()
                    break
                except ValueError:
                    continue
        if not post_date:
            errors.append(f"Row {row_idx}: could not parse date '{raw_date}'")
            continue

        def cell_val(col):
            if col is None:
                return ""
            v = ws.cell(row_idx, col).value
            return str(v).strip() if v else ""

        raw_type = cell_val(col_type) or "Static"
        # Normalize post type
        type_map = {
            "static": "Static", "post": "Static", "reel": "Reel",
            "carousel": "Carousel", "story": "Story", "ugc": "UGC"
        }
        post_type = type_map.get(raw_type.lower(), raw_type)

        raw_platforms = cell_val(col_platform).lower()
        platforms = []
        if "instagram" in raw_platforms:
            platforms.append("instagram")
        if "facebook" in raw_platforms:
            platforms.append("facebook")
        if "linkedin" in raw_platforms:
            platforms.append("linkedin")
        if not platforms:
            platforms = ["instagram"]

        db.add(ContentItem(
            client_id=client.id,
            post_date=post_date,
            post_type=post_type,
            platforms=json.dumps(platforms),
            topic=cell_val(col_topic),
            cover_text=cell_val(col_cover),
            image_text=cell_val(col_image),
            caption=cell_val(col_caption),
            hashtags=json.dumps([]),
            reference_note=cell_val(col_ref),
            status="generated",
            is_on_demand=False,
        ))
        imported += 1

    db.commit()
    return RedirectResponse(
        url=f"/client/{client_id}?imported={imported}",
        status_code=303,
    )


# ── App startup / shutdown ────────────────────────────────────────────────────
from contextlib import asynccontextmanager
from app.services.scheduler import start_scheduler
from app.models.client import LinkedInToken
import secrets

@asynccontextmanager
async def lifespan(app):
    Base.metadata.create_all(bind=engine)
    start_scheduler()
    yield

app.router.lifespan_context = lifespan


# ── LinkedIn OAuth ─────────────────────────────────────────────────────────────
from app.services.linkedin import (
    get_auth_url, exchange_code_for_token, get_linkedin_profile
)
from starlette.middleware.sessions import SessionMiddleware

SECRET_KEY = os.getenv("APP_SECRET_KEY", "influz-studio-dev-secret-2026")
app.add_middleware(SessionMiddleware, secret_key=SECRET_KEY)


@app.get("/auth/linkedin/start/{client_id}")
def linkedin_auth_start(client_id: int, request: Request):
    """Redirect user to LinkedIn OAuth."""
    state = f"{client_id}:{secrets.token_hex(8)}"
    request.session["oauth_state"] = state
    return RedirectResponse(url=get_auth_url(state))


@app.get("/auth/linkedin/callback")
def linkedin_auth_callback(
    request: Request,
    code: str = None,
    state: str = None,
    error: str = None,
    db: Session = Depends(get_db),
):
    if error or not code:
        return HTMLResponse(f"<p style='color:red'>LinkedIn auth failed: {error}</p>")

    stored_state = request.session.get("oauth_state", "")
    if not state or not stored_state.startswith(state.split(":")[0]):
        return HTMLResponse("<p style='color:red'>Invalid OAuth state.</p>")

    client_id = int(stored_state.split(":")[0])

    token_data = exchange_code_for_token(code)
    access_token = token_data.get("access_token", "")

    profile = get_linkedin_profile(access_token)
    person_urn = profile.get("sub", "")
    name = profile.get("name", "")

    existing = db.query(LinkedInToken).filter(LinkedInToken.client_id == client_id).first()
    if existing:
        existing.access_token = access_token
        existing.person_urn = person_urn
        existing.name = name
    else:
        db.add(LinkedInToken(
            client_id=client_id,
            access_token=access_token,
            person_urn=person_urn,
            name=name,
        ))
    db.commit()
    return RedirectResponse(url=f"/client/{client_id}?linkedin=connected")


# ── Creative generation ────────────────────────────────────────────────────────
from app.services.creative import generate_static_creative, generate_carousel_creatives


@app.post("/content/{item_id}/generate-creative")
def generate_creative(item_id: int, db: Session = Depends(get_db)):
    """Generate creative — uses photo (client upload or Unsplash) with text overlay."""
    from app.services.photo_creative import generate_photo_creative
    item = db.query(ContentItem).filter(ContentItem.id == item_id).first()
    if not item:
        raise HTTPException(status_code=404, detail="Not found")
    client = db.query(Client).filter(Client.id == item.client_id).first()

    try:
        if item.post_type == "Carousel":
            # Carousel: generate multiple photo slides
            from app.services.creative import generate_carousel_creatives
            paths = generate_carousel_creatives(
                item.id, item.cover_text or item.topic,
                item.image_text or item.topic, item.post_type)
        else:
            client_photo = None
            try:
                client_photo = item.client_photo_path or None
            except Exception:
                pass
            if not client_photo and item.reference_note and item.reference_note.startswith("PHOTO:"):
                client_photo = item.reference_note.replace("PHOTO:", "").strip()
            path = generate_photo_creative(
                item_id=item.id,
                cover_text=item.cover_text or item.topic,
                image_text=item.image_text or "",
                post_type=item.post_type,
                topic=item.topic,
                niche=client.niche if client else "",
                client_photo_path=client_photo,
                business_name=client.business_name if client else "Influz Studio",
                website="influzstudio.netlify.app",
            )
            paths = [path]

        item.creative_paths = json.dumps(paths)
        item.status = "creative_ready"
        db.commit()
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Creative generation failed: {e}")

    return RedirectResponse(url=f"/client/{item.client_id}/creatives", status_code=303)


@app.post("/content/{item_id}/upload-photo")
async def upload_photo(
    item_id: int,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
):
    item = db.query(ContentItem).filter(ContentItem.id == item_id).first()
    if not item:
        raise HTTPException(status_code=404, detail="Not found")

    client_id = item.client_id

    photos_dir = Path("/tmp/client_photos")
    photos_dir.mkdir(parents=True, exist_ok=True)
    photo_path = photos_dir / f"item_{item_id}_client.jpg"

    contents = await file.read()
    with open(photo_path, "wb") as f:
        f.write(contents)

    # Use raw SQL to update — avoids ORM column issues
    from sqlalchemy import text as _t
    with engine.connect() as conn:
        # Try adding columns if missing
        for col, defn in [
            ("client_photo_path", "VARCHAR DEFAULT ''"),
            ("creative_paths", "TEXT DEFAULT '[]'"),
        ]:
            try:
                conn.execute(_t(f"ALTER TABLE content_items ADD COLUMN {col} {defn}"))
                conn.commit()
            except Exception:
                pass
        # Update the record directly
        conn.execute(_t(
            "UPDATE content_items SET status='approved', "
            "reference_note=:rn, creative_paths='[]' WHERE id=:id"
        ), {"rn": f"PHOTO:{photo_path}", "id": item_id})
        conn.commit()

    return RedirectResponse(
        url=f"/client/{client_id}/creatives", status_code=303
    )


@app.post("/content/{item_id}/approve-creative")
def approve_creative(item_id: int, db: Session = Depends(get_db)):
    """
    Approve creative → auto-publish if post_date is today or past,
    otherwise mark as creative_approved (scheduled — will auto-post on the date).
    """
    from app.services.linkedin import post_to_linkedin
    item = db.query(ContentItem).filter(ContentItem.id == item_id).first()
    if not item:
        raise HTTPException(status_code=404, detail="Not found")

    li_token = db.query(LinkedInToken).filter(LinkedInToken.client_id == item.client_id).first()
    today = date.today()

    # If post_date is today or past → publish immediately
    if item.post_date <= today and li_token:
        platforms = json.loads(item.platforms or "[]")
        creative_paths = json.loads(item.creative_paths or "[]")

        # Regenerate if files missing (ephemeral storage)
        full_paths = []
        for p in creative_paths:
            full = f"/tmp/{p}"
            if not os.path.exists(full):
                try:
                    if item.post_type == "Carousel":
                        generate_carousel_creatives(item.id, item.cover_text or item.topic,
                                                    item.image_text or item.topic, item.post_type)
                    else:
                        generate_static_creative(item.id, item.cover_text or item.topic,
                                                 item.image_text or "", item.post_type, item.topic)
                except Exception:
                    pass
            if os.path.exists(full):
                full_paths.append(full)

        posted_to = []
        if "linkedin" in platforms:
            try:
                try:
                    post_to_linkedin(
                        access_token=li_token.access_token,
                        person_urn=li_token.person_urn,
                        caption=item.caption,
                        image_paths=full_paths or None,
                    )
                except Exception:
                    post_to_linkedin(
                        access_token=li_token.access_token,
                        person_urn=li_token.person_urn,
                        caption=item.caption,
                        image_paths=None,
                    )
                posted_to.append("linkedin")
            except Exception:
                pass  # Mark as approved anyway, scheduler will retry

        if posted_to:
            item.status = "posted"
            item.posted_at = today.isoformat()
            item.posted_to = json.dumps(posted_to)
        else:
            item.status = "creative_approved"
    else:
        # Future date → mark as scheduled, scheduler posts it on the right day
        item.status = "creative_approved"

    db.commit()
    return RedirectResponse(url=f"/client/{item.client_id}/creatives", status_code=303)


@app.post("/content/{item_id}/reject-creative")
def reject_creative(item_id: int, db: Session = Depends(get_db)):
    """Reject creative — sends post back to approved status for regeneration."""
    item = db.query(ContentItem).filter(ContentItem.id == item_id).first()
    if not item:
        raise HTTPException(status_code=404, detail="Not found")
    item.status = "approved"
    item.creative_paths = "[]"
    db.commit()
    return RedirectResponse(
        url=f"/client/{item.client_id}/creatives", status_code=303
    )


@app.post("/content/{item_id}/publish-now")
def publish_now(item_id: int, db: Session = Depends(get_db)):
    """Manually trigger immediate publishing for a creative_approved post."""
    from app.services.linkedin import post_to_linkedin
    item = db.query(ContentItem).filter(ContentItem.id == item_id).first()
    if not item:
        raise HTTPException(status_code=404, detail="Not found")

    client = db.query(Client).filter(Client.id == item.client_id).first()
    li_token = db.query(LinkedInToken).filter(LinkedInToken.client_id == item.client_id).first()

    if not li_token:
        raise HTTPException(status_code=400, detail="LinkedIn not connected for this client")

    platforms = json.loads(item.platforms or "[]")
    creative_paths = json.loads(item.creative_paths or "[]")

    # Regenerate creative if file doesn't exist (Render ephemeral filesystem)
    full_paths = []
    for p in creative_paths:
        full = f"/tmp/{p}"
        if not os.path.exists(full):
            try:
                if item.post_type == "Carousel":
                    generate_carousel_creatives(item.id, item.cover_text or item.topic,
                                                item.image_text or item.topic, item.post_type)
                else:
                    generate_static_creative(item.id, item.cover_text or item.topic,
                                             item.image_text or "", item.post_type, item.topic)
            except Exception:
                pass
        if os.path.exists(full):
            full_paths.append(full)

    posted_to = []
    if "linkedin" in platforms:
        try:
            # Try with image first, fall back to text-only if image upload fails
            try:
                post_to_linkedin(
                    access_token=li_token.access_token,
                    person_urn=li_token.person_urn,
                    caption=item.caption,
                    image_paths=full_paths or None,
                )
            except Exception:
                # Fallback: post text only
                post_to_linkedin(
                    access_token=li_token.access_token,
                    person_urn=li_token.person_urn,
                    caption=item.caption,
                    image_paths=None,
                )
            posted_to.append("linkedin")
        except Exception as e:
            raise HTTPException(status_code=500, detail=f"LinkedIn posting failed: {e}")

    if posted_to:
        item.status = "posted"
        item.posted_at = date.today().isoformat()
        item.posted_to = json.dumps(posted_to)
        db.commit()

    return RedirectResponse(
        url=f"/client/{item.client_id}/creatives", status_code=303
    )


# ── Creatives review page ──────────────────────────────────────────────────────
from pathlib import Path as _Path


@app.get("/client/{client_id}/creatives", response_class=HTMLResponse)
def creatives_page(request: Request, client_id: int, db: Session = Depends(get_db)):
    client = db.query(Client).filter(Client.id == client_id).first()
    if not client:
        raise HTTPException(status_code=404, detail="Not found")

    li_token = db.query(LinkedInToken).filter(LinkedInToken.client_id == client_id).first()

    items = (
        db.query(ContentItem)
        .filter(ContentItem.client_id == client_id)
        .order_by(ContentItem.post_date)
        .all()
    )

    # Enrich items with parsed creative paths
    enriched = []
    for item in items:
        paths = json.loads(item.creative_paths or "[]")
        platforms = json.loads(item.platforms or "[]")
        enriched.append({
            "item": item,
            "creative_paths": paths,
            "platforms": platforms,
        })

    return templates.TemplateResponse(
        "creatives.html",
        {
            "request": request,
            "client": client,
            "enriched": enriched,
            "li_token": li_token,
        },
    )
